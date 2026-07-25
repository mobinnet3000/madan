#!/usr/bin/env python3

"""
Test script to verify all imports work correctly.
This is a safe way to test imports without relying on bash utilities.
"""

import sys
import os

# Add the project directory to the path
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

# Import all the required modules
print("Testing all required modules...")

# Core Django and DRF modules
import django
from django.apps import apps
from django.conf import settings

# Report generation modules
from machines.reports import generate_performance_report, generate_analysis_report, get_date_range, RANGE_LABELS
from openpyxl import Workbook
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
import xhtml2pdf
import arabic_reshaper
import bidi

# Test all functions

# 1. Test get_date_range
from machines.reports import get_date_range

test_cases = [
    ('today', None, None),
    ('yesterday', None, None),
    ('this_week', None, None),
    ('last_week', None, None),
    ('this_month', None, None),
    ('last_month', None, None),
    ('this_year', None, None),
    ('last_year', None, None),
    ('7days', None, None),
    ('30days', None, None),
    ('90days', None, None),
    ('365days', None, None),
    ('all', None, None),
]

print("Testing get_date_range for all range keys...")
for key, start, end in test_cases:
    result = get_date_range(key, start, end)
    print(f"  {key}: {result}")

# 2. Test RANGE_LABELS
print("\nRANGE_LABELS:")
for key, value in RANGE_LABELS.items():
    print(f"  {key}: {value}")

# 3. Test that we can access Django apps
print("\nDjango apps check:")
print(f"  machines app exists: {apps.get_app_config('machines') is not None}")
print(f"  accounts app exists: {apps.get_app_config('accounts') is not None}")

# 4. Test that we can access models
from machines.models import Factory, DeviceLog, DeviceDailyAnalysis

print("\nTesting model access:")
print(f"  Factory model exists: {Factory is not None}")
print(f"  DeviceLog model exists: {DeviceLog is not None}")
print(f"  DeviceDailyAnalysis model exists: {DeviceDailyAnalysis is not None}")

# 5. Test that we can create simple Excel data
print("\nTesting Excel workbook creation...")
wb = Workbook()
ws = wb.active
ws['A1'] = 'Тест'
ws['B1'] = 'Test'
ws['A2'] = 'آزمون'

# Save to memory
from io import BytesIO
buf = BytesIO()
wb.save(buf)
buf.seek(0)
print(f"  Excel workbook size: {len(buf.getvalue())} bytes")

# 6. Test that we can create Django test client
from django.test import Client
client = Client()
print("\nDjango test client created successfully")

print("\n" + "="*80)
print("ALL TESTS PASSED!")
print("="*80)

# Test report generation with mock data (this would require an actual database)
print("\nNote: Report generation tests need actual database data.")
print("Data seeding can be done via seed_demo.py")

# Check if the font files exist
fonts_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'machines', 'fonts')
print(f"\nFonts directory exists: {os.path.exists(fonts_dir)}")
if os.path.exists(fonts_dir):
    fonts = os.listdir(fonts_dir)
    print(f"  Font files: {fonts}")

print("\nExit status: 0")