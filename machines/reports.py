import os, io, calendar, locale
from datetime import date, timedelta, datetime
from django.conf import settings
from django.db.models import Sum, Avg, Count, Q, Min, Max
from django.utils import timezone
from collections import OrderedDict

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.units import mm, cm
from reportlab.lib import colors
from reportlab.platypus import (SimpleDocTemplate, Table, TableStyle, Paragraph,
                                Spacer, Image, PageBreak, KeepTogether)
from reportlab.lib.styles import ParagraphStyle
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont

from .models import Factory, ProductionLine, Device, DeviceLog
from .jalali import to_jalali, to_jalali_full, weekday_fa

# Arabic/Persian text reshaping for correct glyph rendering in Reportlab
import arabic_reshaper
from bidi.algorithm import get_display

FONTS_DIR = os.path.join(os.path.dirname(__file__), 'fonts')

PERSIAN_MONTHS = ['فروردین', 'اردیبهشت', 'خرداد', 'تیر', 'مرداد', 'شهریور',
                  'مهر', 'آبان', 'آذر', 'دی', 'بهمن', 'اسفند']

RANGE_LABELS = {
    'today': 'امروز',
    'yesterday': 'دیروز',
    'this_week': 'این هفته',
    'last_week': 'هفته گذشته',
    'this_month': 'این ماه',
    'last_month': 'ماه گذشته',
    'this_year': 'امسال',
    'last_year': 'سال گذشته',
    '7days': '۷ روز گذشته',
    '30days': '۳۰ روز گذشته',
    '90days': '۹۰ روز گذشته',
    '365days': 'یک سال گذشته',
    'all': 'همه',
    'custom': 'سفارشی',
}

# ---------------------------------------------------------------------------
#  Persian text pipeline
# ---------------------------------------------------------------------------
def _fa(text):
    """Reshape + bidi for correct Persian rendering in Reportlab."""
    if not text:
        return ''
    try:
        reshaped = arabic_reshaper.reshape(str(text))
        return get_display(reshaped)
    except Exception:
        return str(text)

def _p(value, decimals=1):
    """Format number to Persian digits."""
    if value is None:
        return '*۰*'
    try:
        v = float(value)
    except (TypeError, ValueError):
        return str(value)
    s = f'{v:,.{decimals}f}'.translate(str.maketrans('0123456789-,.', '۰۱۲۳۴۵۶۷۸۹-،.'))
    return s

def _factory_header(factory):
    return [
        f'شرکت: {factory.name if factory else ""}',
        f'آدرس: {factory.address if factory and factory.address else ""}',
    ]

# ---------------------------------------------------------------------------
#  Font registration (robust)
# ---------------------------------------------------------------------------
def _register_fonts():
    candidates = {
        'VazirBold':   os.path.join(FONTS_DIR, 'Vazirmatn-FD-Bold.ttf'),
        'Vazir':       os.path.join(FONTS_DIR, 'Vazirmatn-FD-Regular.ttf'),
        'VazirMedium': os.path.join(FONTS_DIR, 'Vazirmatn-FD-Medium.ttf'),
    }
    for name, path in candidates.items():
        if os.path.exists(path):
            pdfmetrics.registerFont(TTFont(name, path))

_register_fonts()

# Base colours
C_PRIMARY   = colors.HexColor('#1a365d')   # dark navy
C_ACCENT    = colors.HexColor('#2563eb')    # blue
C_HEADER_BG = colors.HexColor('#1e3a5f')
C_ROW_EVEN  = colors.HexColor('#f1f5f9')
C_ROW_ODD   = colors.white
C_BORDER    = colors.HexColor('#cbd5e1')
C_TEXT_DARK = colors.HexColor('#1e293b')
C_TEXT_MUTED= colors.HexColor('#64748b')
C_WHITE     = colors.white
C_GOLD      = colors.HexColor('#f59e0b')
C_TOTAL_BG  = colors.HexColor('#e2e8f0')

STYLES = {
    'title': ParagraphStyle('TitleFa', fontName='VazirBold', fontSize=18, alignment=1,
                            textColor=C_PRIMARY, spaceAfter=2),
    'subtitle': ParagraphStyle('SubFa', fontName='VazirMedium', fontSize=10, alignment=1,
                               textColor=C_TEXT_MUTED, spaceAfter=2),
    'header_section': ParagraphStyle('HeadSec', fontName='VazirBold', fontSize=11,
                                     textColor=C_PRIMARY, spaceAfter=4, spaceBefore=6),
    'cell_left': ParagraphStyle('CellLt', fontName='Vazir', fontSize=8,
                                leading=11, alignment=0, textColor=C_TEXT_DARK),
    'cell_center': ParagraphStyle('CellCt', fontName='Vazir', fontSize=8,
                                  leading=11, alignment=1, textColor=C_TEXT_DARK),
    'cell_header': ParagraphStyle('CellHd', fontName='VazirBold', fontSize=9,
                                  leading=12, alignment=1, textColor=C_WHITE),
    'meta': ParagraphStyle('MetaFa', fontName='Vazir', fontSize=8,
                           textColor=C_TEXT_MUTED, spaceAfter=2),
    'total_label': ParagraphStyle('TotLb', fontName='VazirBold', fontSize=9,
                                  alignment=1, textColor=C_PRIMARY),
    'total_value': ParagraphStyle('TotVl', fontName='VazirBold', fontSize=9,
                                  alignment=1, textColor=C_PRIMARY),
}

# --- Date range helpers ---
def get_date_range(range_key, start_date=None, end_date=None):
    today = date.today()
    if range_key == 'today':
        return today, today
    elif range_key == 'yesterday':
        y = today - timedelta(days=1)
        return y, y
    elif range_key == 'this_week':
        start = today - timedelta(days=today.weekday())
        return start, today
    elif range_key == 'last_week':
        start = today - timedelta(days=today.weekday() + 7)
        return start, start + timedelta(days=6)
    elif range_key == 'this_month':
        return today.replace(day=1), today
    elif range_key == 'last_month':
        first = today.replace(day=1) - timedelta(days=1)
        return first.replace(day=1), first
    elif range_key == 'this_year':
        return today.replace(month=1, day=1), today
    elif range_key == 'last_year':
        return today.replace(year=today.year - 1, month=1, day=1), today.replace(year=today.year - 1, month=12, day=31)
    elif range_key == '7days':
        return today - timedelta(days=7), today
    elif range_key == '30days':
        return today - timedelta(days=30), today
    elif range_key == '90days':
        return today - timedelta(days=90), today
    elif range_key == '365days':
        return today - timedelta(days=365), today
    elif range_key == 'all':
        return None, None
    elif range_key == 'custom' and start_date and end_date:
        return start_date, end_date
    return today - timedelta(days=30), today

def _p(value, decimals=1):
    if value is None:
        return '۰'
    return f'{value:,.{decimals}f}'.translate(str.maketrans('0123456789-,.', '۰۱۲۳۴۵۶۷۸۹-،.'))

def _factory_header(factory):
    return [
        f'شرکت: {factory.name}',
        f'آدرس: {factory.address}',
    ]

# ────────────────────────────────── EXCEL ──────────────────────────────────

def _excel_style(ws):
    header_font = Font(name='Vazirmatn', bold=True, color='FFFFFF', size=11)
    header_fill = PatternFill(start_color='1a56db', end_color='1a56db', fill_type='solid')
    header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    return header_font, header_fill, header_align, thin_border

def _write_rows(ws, rows, row_offset=1):
    hf, hfill, halign, border = _excel_style(ws)
    for r_idx, row in enumerate(rows, start=row_offset):
        for c_idx, val in enumerate(row, start=1):
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            cell.border = border
            if r_idx == row_offset:
                cell.font = hf
                cell.fill = hfill
                cell.alignment = halign
            else:
                cell.alignment = Alignment(horizontal='center', vertical='center')

def excel_performance(factory, lines_data, date_from, date_to, date_label, factory_obj=None):
    wb = Workbook()
    ws = wb.active
    ws.title = 'توقفات خط تولید'

    data_rows = []
    data_rows.append(['توقفات خط تولید - ' + (factory.name if factory else 'کل کارخانه‌ها')] + [''] * 7)
    data_rows.append(['بازه: ' + date_label] + [''] * 7)
    data_rows.append([])
    headers = ['ردیف', 'خط تولید', 'تعداد ثبت', 'ساعت کارکرد', 'ساعت توقف',
               'ورودی (تن)', 'محصول (تن)', 'باطله (تن)', 'بازدهی (%)']

    data_rows.append(headers)
    for i, (line_name, line_id, stats) in enumerate(lines_data, 1):
        n = stats.get('count', 0)
        run = stats.get('runtime', 0)
        down = stats.get('downtime', 0)
        feed = stats.get('feed', 0)
        prod = stats.get('product', 0)
        tail = stats.get('tailing', 0)
        eff = stats.get('efficiency', 0)
        data_rows.append([i, line_name, n, run, down, feed, prod, tail, eff])

    title_font = Font(name='Vazirmatn', bold=True, size=14)
    subtitle_font = Font(name='Vazirmatn', size=10)

    for r_idx, row in enumerate(data_rows, 1):
        for c_idx, val in enumerate(row, 1):
            cell = ws.cell(row=r_idx, column=c_idx)
            cell.value = val
            if r_idx == 1:
                cell.font = title_font
                cell.alignment = Alignment(horizontal='center')
            elif r_idx == 2:
                cell.font = subtitle_font
                cell.alignment = Alignment(horizontal='center')

    _write_rows(ws, data_rows[3:], row_offset=4)

    for col in range(1, 10):
        ws.column_dimensions[get_column_letter(col)].width = 14

    total_runtime = sum(s.get('runtime', 0) for _, _, s in lines_data)
    total_downtime = sum(s.get('downtime', 0) for _, _, s in lines_data)
    total_feed = sum(s.get('feed', 0) for _, _, s in lines_data)
    total_product = sum(s.get('product', 0) for _, _, s in lines_data)
    total_tailing = sum(s.get('tailing', 0) for _, _, s in lines_data)

    last_row = len(data_rows) + 3
    ws.cell(row=last_row, column=1, value='جمع کل').font = Font(name='Vazirmatn', bold=True, size=11)
    ws.cell(row=last_row, column=2, value='-')
    ws.cell(row=last_row, column=3, value='-')
    ws.cell(row=last_row, column=4, value=total_runtime).font = Font(name='Vazirmatn', bold=True, size=11)
    ws.cell(row=last_row, column=5, value=total_downtime).font = Font(name='Vazirmatn', bold=True, size=11)
    ws.cell(row=last_row, column=6, value=total_feed).font = Font(name='Vazirmatn', bold=True, size=11)
    ws.cell(row=last_row, column=7, value=total_product).font = Font(name='Vazirmatn', bold=True, size=11)
    ws.cell(row=last_row, column=8, value=total_tailing).font = Font(name='Vazirmatn', bold=True, size=11)
    eff_all = (total_product / total_feed * 100) if total_feed else 0
    ws.cell(row=last_row, column=9, value=round(eff_all, 1)).font = Font(name='Vazirmatn', bold=True, size=11)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf

# ────────────────────────────────── PDF ──────────────────────────────────

# ---------- helpers ----------
def _pdf_header_footer(canvas, doc, title, factory=None, date_label=''):
    canvas.saveState()
    canvas.setFont('Vazir', 7)
    canvas.setFillColor(C_TEXT_MUTED)
    w, h = A4 if doc.pagesize == A4 else landscape(A4)
    today_str = to_jalali(date.today())
    canvas.drawRightString(w - 20, 15, _fa(f'صفحه {doc.page}'))
    canvas.drawRightString(w - 20, 10, _fa(f'تاریخ چاپ: {today_str}'))
    if factory:
        canvas.drawString(20, 10, _fa(f'{factory.name} | {title}'))
    canvas.restoreState()

def _p_cell(text, style='cell_center'):
    return Paragraph(_fa(str(text)), STYLES[style])

def _make_table(data, col_widths=None, header_rows=1):
    t = Table(data, colWidths=col_widths, repeatRows=header_rows)
    cmds = [
        ('FONTNAME', (0, 0), (-1, header_rows - 1), 'VazirBold'),
        ('FONTNAME', (0, header_rows), (-1, -1), 'Vazir'),
        ('FONTSIZE', (0, 0), (-1, -1), 8),
        ('BACKGROUND', (0, 0), (-1, header_rows - 1), C_HEADER_BG),
        ('TEXTCOLOR', (0, 0), (-1, header_rows - 1), C_WHITE),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('GRID', (0, 0), (-1, -1), 0.5, C_BORDER),
        ('TOPPADDING', (0, 0), (-1, -1), 3),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 3),
        ('LEFTPADDING', (0, 0), (-1, -1), 4),
        ('RIGHTPADDING', (0, 0), (-1, -1), 4),
    ]
    # Alternate row colours
    nrows = len(data) - header_rows
    for i in range(header_rows, len(data)):
        bg = C_ROW_EVEN if (i - header_rows) % 2 == 0 else C_ROW_ODD
        cmds.append(('BACKGROUND', (0, i), (-1, i), bg))
    # Apply header row background
    t.setStyle(TableStyle(cmds))
    return t

# ---------- performance ----------
def pdf_performance(factory, lines_data, date_from, date_to, date_label, factory_obj=None):
    from reportlab.platypus import Table as T
    buf = io.BytesIO()
    doc = SimpleDocTemplate(
        buf, pagesize=A4,
        rightMargin=14*mm, leftMargin=14*mm,
        topMargin=18*mm, bottomMargin=18*mm,
    )

    elements = []

    # ── title block ──
    elements.append(Paragraph(_fa('توقفات خطوط تولید'), STYLES['title']))
    elements.append(Spacer(1, 3))

    info_lines = []
    if factory:
        info_lines.append(Paragraph(_fa(f'شرکت: {factory.name}'), STYLES['subtitle']))
        if factory.address:
            info_lines.append(Paragraph(_fa(f'آدرس: {factory.address}'), STYLES['meta']))
    info_lines.append(Paragraph(_fa(f'بازه گزارش: {date_label}'), STYLES['meta']))
    for l in info_lines:
        elements.append(l)
    elements.append(Spacer(1, 8))

    # ── main table ──
    hdr = [_fa('ردیف'), _fa('خط تولید'), _fa('تعداد'), _fa('کارکرد'),
           _fa('توقف'), _fa('خوراک (تن)'), _fa('محصول (تن)'),
           _fa('باطله (تن)'), _fa('بازدهی')]
    cw = [28, 95, 42, 50, 45, 55, 55, 55, 48]
    rows = [hdr]
    for i, (ln, lid, st) in enumerate(lines_data, 1):
        rows.append([
            _p_cell(str(i)), _p_cell(ln),
            _p_cell(_p(st.get('count', 0), 0)),
            _p_cell(_p(st.get('runtime', 0))),
            _p_cell(_p(st.get('downtime', 0))),
            _p_cell(_p(st.get('feed', 0))),
            _p_cell(_p(st.get('product', 0))),
            _p_cell(_p(st.get('tailing', 0))),
            _p_cell(f"{_p(st.get('efficiency', 0))}%"),
        ])

    elements.append(_make_table(rows, col_widths=cw))

    # ── summary row ──
    if len(lines_data) > 1:
        elements.append(Spacer(1, 10))
        tr = sum(s.get('runtime', 0) for _, _, s in lines_data)
        td = sum(s.get('downtime', 0) for _, _, s in lines_data)
        tf = sum(s.get('feed', 0) for _, _, s in lines_data)
        tp = sum(s.get('product', 0) for _, _, s in lines_data)
        tt = sum(s.get('tailing', 0) for _, _, s in lines_data)
        ea = (tp / tf * 100) if tf else 0

        srow = [
            _p_cell(''), _p_cell(_fa('جمع کل'), 'total_label'),
            _p_cell(_p(tp, 0), 'total_label'),
            _p_cell(_p(tr), 'total_value'),
            _p_cell(_p(td), 'total_value'),
            _p_cell(_p(tf), 'total_value'),
            _p_cell(_p(tp), 'total_value'),
            _p_cell(_p(tt), 'total_value'),
            _p_cell(f"{_p(ea)}%", 'total_value'),
        ]

        st = Table([srow], colWidths=cw)
        st.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, -1), C_TOTAL_BG),
            ('FONTNAME', (0, 0), (-1, -1), 'VazirBold'),
            ('FONTSIZE', (0, 0), (-1, -1), 9),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('GRID', (0, 0), (-1, -1), 0.5, C_BORDER),
            ('TOPPADDING', (0, 0), (-1, -1), 4),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
        ]))
        elements.append(st)

    doc.build(
        elements,
        onFirstPage=lambda c, d: _pdf_header_footer(c, d, 'توقفات خط تولید', factory, date_label),
        onLaterPages=lambda c, d: _pdf_header_footer(c, d, 'توقفات خط تولید', factory, date_label),
    )
    buf.seek(0)
    return buf


# ─────────────────────────────── QUERIES ───────────────────────────────

def performance_data(factory, date_from, date_to):
    qs = DeviceLog.objects.all()
    if factory:
        qs = qs.filter(line__factory=factory)
    if date_from:
        qs = qs.filter(date__gte=date_from)
    if date_to:
        qs = qs.filter(date__lte=date_to)

    lines_qs = qs.values('line_id', 'line__name').annotate(
        count=Count('id'),
        runtime=Sum('runtime_hours'),
        downtime=Sum('downtime_hours'),
        feed=Sum('feed_tonnage'),
        product=Sum('product_tonnage'),
        tailing=Sum('tailing_tonnage'),
    ).order_by('-line__name')

    lines_data = []
    for l in lines_qs:
        eff = (l['product'] / l['feed'] * 100) if l['feed'] and l['feed'] > 0 else 0
        stats = {
            'count': l['count'],
            'runtime': float(l['runtime'] or 0),
            'downtime': float(l['downtime'] or 0),
            'feed': float(l['feed'] or 0),
            'product': float(l['product'] or 0),
            'tailing': float(l['tailing'] or 0),
            'efficiency': round(eff, 1),
        }
        lines_data.append((l['line__name'], l['line_id'], stats))

    return lines_data


# ─────────────────────────────── MAIN ───────────────────────────────

def _date_label(range_key, d_from, d_to):
    """ساخت برچسب بازه گزارش با تاریخ شمسی + روز هفته."""
    if range_key == 'custom' and d_from and d_to:
        return f'{to_jalali_full(d_from)} ({weekday_fa(d_from)}) تا {to_jalali_full(d_to)} ({weekday_fa(d_to)})'
    if d_from and d_to:
        if d_from == d_to:
            return f'{to_jalali_full(d_from)} - {weekday_fa(d_from)}'
        return f'{to_jalali_full(d_from)} ({weekday_fa(d_from)}) تا {to_jalali_full(d_to)} ({weekday_fa(d_to)})'
    return 'همه موارد'


def generate_performance_report(factory_id, range_key, date_from=None, date_to=None, fmt='excel'):
    factory = Factory.objects.filter(id=factory_id).first() if factory_id else None
    d_from, d_to = get_date_range(range_key, date_from, date_to)
    date_label = _date_label(range_key, d_from, d_to)

    lines_data = performance_data(factory, d_from, d_to)

    if fmt == 'excel':
        buf = excel_performance(factory, lines_data, d_from, d_to, date_label)
        return buf, 'xlsx'
    else:
        buf = pdf_performance(factory, lines_data, d_from, d_to, date_label)
        return buf, 'pdf'
